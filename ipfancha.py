import openpyxl
import requests
import time
import re
import numpy
requests.packages.urllib3.disable_warnings()
requests.adapters.DEFAULT_RETRIES =5
book=openpyxl.load_workbook('D:\\安服脚本\\output\\fofa导出结果1639923830827.xlsx')
# book1=openpyxl.workbook('D:\安服脚本\output\fofa导出结果1639318511164.xlsx',encodings='utf-8')
# book1.active
# book1.create_sheet('ip',0)
sheet=book['查询结果']
header = {
    'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.93 Safari/537.36',
}
proxy={
    'http':'http://127.0.0.1:8080',
    'https':'https://127.0.0.1:8080'
}
pat=re.compile(r'<li><span class="date">.*</span><a href="/.*/" target="_blank">.*</a></li>')
pat2=re.compile(r'"/.*/"')
sessions=requests.session()
i=1
while i<=sheet.max_row:
    i=i+1
    tar=sheet.cell(row=i,column=4).value
    try:
        tar1 = 'http://' + sheet.cell(row=i, column=1).value
        rres = sessions.get(url=tar1, headers=header, verify=False,allow_redirects=False,proxies=proxy)
    except requests.exceptions.ConnectionError:
        continue
    if rres.status_code!=200 and rres.status_code!=302:
        print('目标站点不可访问')
        continue
    url = 'https://site.ip138.com/{0}'.format(tar)
    try:
        res = sessions.get(url=url, headers=header, verify=False)
        if '暂无结果' not in res.text:
            print(tar)
            resu = numpy.array(re.findall(pat, res.text))
            for n in resu:
                resu1 = str(re.findall(pat2, str(n)))
                resu2=resu1.replace('"','')
                resu3 =resu2.replace('/','')
                resu4 = resu3.replace('\'', '')
                resu5 = resu4.replace('[', '')
                resu6 = resu5.replace(']', '')
                print(resu6)
        if 'Burp Suite Professional' in res.text:
            exit()
        else:
            print('此ip无域名')
    except (requests.exceptions.ConnectionError,NameError):
        print('连接失败\n')
        i = i - 1
        time.sleep(100)
        continue
    time.sleep(0.5)